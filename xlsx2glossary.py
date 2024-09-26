"""
Python Script to convert a list of terms listed in a xlsx sheet to an XML glossary conforming to the Moodle glossary template


"""
import lxml.etree as et
from openpyxl import load_workbook

# Load the spreadsheet with the terms
wb = load_workbook('terms.xlsx')

# Select the Terms sheet with the data
ws = wb['Terms']

# Find the column names (assumed to be on the first (header) line)
ColNames = {}
Current = 0
for COL in ws.iter_cols(1, ws.max_column):
    ColNames[COL[0].value] = Current
    Current += 1

# Set the term and definition columns
term_column = ColNames['Term']
definition_column = ColNames['Definition']

# parse the template
tree = et.parse('GlossaryTemplate.xml')

# get the info element
info = tree.findall('.//INFO')[0]

# Select the entries element
q = tree.findall('.//ENTRIES')[0]

# Iterate over the terms
for row in ws.iter_rows():
    term = row[term_column].value
    definition = row[definition_column].value

    # Looad the template of a glossary entry.
    entry = et.XML("<ENTRY>"
                   "<CONCEPT>Term String</CONCEPT>"
                   "<DEFINITION>Definition String</DEFINITION>"
                   "<FORMAT>0</FORMAT>"
                   "<USEDYNALINK>1</USEDYNALINK>"
                   "<CASESENSITIVE>0</CASESENSITIVE>"
                   "<FULLMATCH>1</FULLMATCH>"
                   "<TEACHERENTRY>1</TEACHERENTRY>"
                   "</ENTRY>")

    # update the entry
    for child in entry:
        if child.tag == 'CONCEPT':
            child.text = term
        if child.tag == 'DEFINITION':
            if definition:
                child.text = definition
            else:
                child.text = 'TO BE DEFINED'

    # add the entry to the collection
    q.append(entry)

# Append the entries to the template
info.append(q)

# Write the final output to a xml file named output.xml
tree.write('output.xml', pretty_print=True, xml_declaration=True, encoding="utf-8")

# Load the output.xml into your moodle instance
